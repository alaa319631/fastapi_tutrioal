from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
import os
from openpyxl import Workbook, load_workbook

app = FastAPI()
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(
  CORSMiddleware,
  allow_origins=["*"],  # يسمح بالوصول من أي مصدر. قم بتقييد هذا في الإنتاج
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)
templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
async def read_form(request: Request):
    return templates.TemplateResponse("form.html", {"request": request})

@app.post("/submit/")
async def handle_form(name: str = Form(...), email: str = Form(...)):
    file_path = 'data.xlsx'
    file_exists = os.path.isfile(file_path)

    if file_exists:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # إضافة العناوين للأعمدة إذا كان الملف جديد
        sheet['A1'] = "Name"
        sheet['B1'] = "Email"

    # الحصول على الصف الأول المتاح بعد العناوين
    next_row = sheet.max_row + 1

    # كتابة الاسم والبريد الإلكتروني في الصف التالي
    sheet[f'A{next_row}'] = name
    sheet[f'B{next_row}'] = email

    # حفظ الملف
    workbook.save(file_path)

    return {"message": "Data stored successfully"}

